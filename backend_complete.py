import os
import sys
import asyncio
import logging
import shutil
import time
import json
import re
import subprocess
from typing import List, Optional, Dict, Any, Generator
from datetime import datetime, timedelta, timezone
from contextlib import contextmanager
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, status, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import JSONResponse
import uvicorn
import psycopg
from psycopg.rows import dict_row
from passlib.context import CryptContext
from jose import JWTError, jwt
try:
    import google.generativeai as genai
except ImportError:
    genai = None
    logging.warning("google.generativeai module not found. Gemini LLM will not be available.")
import aiofiles
import httpx
from openpyxl import load_workbook
import PyPDF2
import docx

# --- Globals for lazily loaded modules ---
create_semantic_chunks = None
retriever = None
context_optimizer = None
collection = None

def initialize_services():
    """
    Initializes all slow, blocking services in a separate thread.
    This includes NLTK, sentence-transformers, and ChromaDB.
    """
    global create_semantic_chunks, retriever, context_optimizer, collection
    logger.debug("Initializing background services...")

    # 1. Initialize NLTK and chunking utils
    logger.debug("Loading chunking utilities (and NLTK)...")
    try:
        from chunking_utils import create_semantic_chunks as csc
        create_semantic_chunks = csc
        logger.debug("Chunking utilities loaded successfully.")
    except ImportError:
        logger.warning("chunking_utils.py not found. Falling back to basic chunking.")
        def fallback_chunks(text, **kwargs): return [text]
        create_semantic_chunks = fallback_chunks
    except Exception as e:
        logger.error(f"An unexpected error occurred while loading chunking_utils: {e}")

    # 2. Initialize advanced retrieval models (sentence-transformers)
    logger.debug("Loading advanced retrieval models...")
    try:
        from advanced_retrieval import retriever as r, context_optimizer as co
        retriever = r
        context_optimizer = co
        logger.debug("Advanced retrieval models loaded successfully.")
    except ImportError:
        logger.warning("advanced_retrieval.py not found. Advanced retrieval features will be disabled.")
    except Exception as e:
        logger.error(f"An unexpected error occurred while loading advanced_retrieval: {e}")

    # 3. Initialize ChromaDB
    try:
        if not SKIP_EMBEDDINGS:
            import chromadb
            from chromadb.config import Settings
            chroma_client = chromadb.PersistentClient(path=CHROMA_DB_DIR)
            collection = chroma_client.get_or_create_collection(name="documents")
        else:
            pass
    except Exception as e:
        logger.error(f"ChromaDB initialization failed: {e}")

    # 4. Check LLM Health
    logger.debug("Checking LLM health...")
    if not check_llm_health():
        logger.warning("LLM Health Check Failed during startup. Check configuration.")
    else:
        logger.debug("LLM Health Check successful.")

    logger.debug("Background services initialization complete.")

# --- Configuration ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Database Config
DB_NAME = os.getenv("DB_NAME", "mvsr_rag")
DB_USER = os.getenv("DB_USER", "postgres")
DB_PASSWORD = os.getenv("DB_PASSWORD", "password") # Default password, change as needed
DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")

# Auth Config
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-here")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# LLM Config
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY and genai:
    genai.configure(api_key=GEMINI_API_KEY)

LLM_PROVIDER = os.getenv("LLM_PROVIDER", "ollama") # or 'ollama'

# Paths
UPLOAD_DIR = os.path.join(os.getcwd(), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
CHROMA_DB_DIR = os.path.join(os.getcwd(), "chroma_db")

# Skip Embeddings Flag (for debugging)
SKIP_EMBEDDINGS = os.getenv("SKIP_EMBEDDINGS", "0") == "1"

# --- Security ---
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# --- Database ---
async def get_db():
    conn_info = f"dbname='{DB_NAME}' user='{DB_USER}' password='{DB_PASSWORD}' host='{DB_HOST}' port='{DB_PORT}'"
    aconn = None
    try:
        for attempt in range(3):
            try:
                aconn = await psycopg.AsyncConnection.connect(conn_info)
                break
            except psycopg.OperationalError as e:
                if attempt == 2:
                    logger.error(f"Database connection failed after 3 attempts: {e}")
                    logger.warning("Falling back to mock data mode. All database operations will be simulated.")
                    logger.warning("Please check your database credentials (DB_NAME, DB_USER, DB_PASSWORD, DB_HOST, DB_PORT) and ensure the database server is running.")
                await asyncio.sleep(1)
        
        if aconn:
            yield aconn
        else:
            yield None
    finally:
        if aconn:
            await aconn.close()

async def get_current_user(token: str = Depends(oauth2_scheme), db: psycopg.AsyncConnection = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    if db:
        async with db.cursor(row_factory=dict_row) as acur:
            await acur.execute("SELECT * FROM users WHERE username = %s", (username,))
            user = await acur.fetchone()
            if user is None:
                raise credentials_exception
            return user
    else:
        # Mock user for offline mode
        if username == "admin":
            return {"username": "admin", "role": "admin"}
        raise credentials_exception

# --- LLM & RAG ---
def check_llm_health():
    if LLM_PROVIDER.lower() == "ollama":
        try:
            # Check local ollama
            response = httpx.get("http://localhost:11434/api/tags")
            return response.status_code == 200
        except:
            return False
    else:
        if not GEMINI_API_KEY:
            return False
        try:
            model = genai.GenerativeModel('gemini-1.5-flash')
            # Simple generation check
            response = model.generate_content("Ping")
            return True
        except Exception as e:
            logger.error(f"Gemini Health Check Failed: {e}")
            return False

def generate_answer_with_llm(context: str, query: str, history: List[Dict[str, str]] = None) -> str:
    if not create_semantic_chunks:
        # Services are not loaded yet
        return "The system is still initializing. Please try again in a moment."

    if not context or not context.strip():
        # If context is empty, provide a direct, helpful response instead of asking for clarification.
        return "I could not find any information about that in the uploaded documents. Please ensure the relevant document has been uploaded and processed correctly."

    history_str = ""
    if history:
        for msg in history:
            role = "User" if msg['role'] == 'user' else "Assistant"
            history_str += f"{role}: {msg['content']}\n"

    system_prompt = f"""You are an expert assistant for a college's research database.
    Your primary goal is to provide accurate, concise, and relevant answers based *only* on the provided text context.

    **Conversation History:**
    ---
    {history_str}
    ---

    **Instructions:**
    1.  **Strictly Adhere to Context:** Your answer MUST be based exclusively on the text provided in the 'Context' section. Do not use any outside knowledge.
    2.  **Use Conversation History:** Pay attention to the 'Conversation History' to understand follow-up questions.
    3.  **Answer the Query Directly:** If the context contains the information to answer the user's query, provide a direct and comprehensive answer. Synthesize the information into a clear response.
    4.  **Handle Missing Information:** If the context does NOT contain the information to answer the query, you MUST state that you could not find the information in the provided documents.
    5.  **No Clarifying Questions:** Do not ask the user clarifying questions. If the query is ambiguous and the context is unhelpful, state that you cannot answer based on the provided information.
    6.  **Formatting:** Use bullet points for lists where appropriate.
    7.  **Be Concise:** Avoid filler and unnecessary preamble.

    Context:
    ---
    {context}
    ---

    User Query: {query}

    Answer:
    """

    try:
        if LLM_PROVIDER.lower() == "ollama":
            # Call Ollama
            response = httpx.post("http://localhost:11434/api/generate", json={
                "model": "llama3.2:3b",
                "prompt": system_prompt,
                "stream": False
            }, timeout=120.0)
            if response.status_code == 200:
                return response.json().get("response", "Error generating response")
            return "Error calling Ollama"
        else:
            # Call Gemini
            model = genai.GenerativeModel('gemini-1.5-flash')
            response = model.generate_content(system_prompt)
            return response.text
    except Exception as e:
        logger.error(f"LLM Generation Error: {e}")
        return f"I encountered an error while processing your request: {str(e)}"

# --- File Processing ---
def extract_text_from_pdf(file_path):
    text = ""
    try:
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() + "\n"
    except Exception as e:
        logger.error(f"PDF extraction error: {e}")
    return text

def extract_text_from_docx(file_path):
    text = ""
    try:
        doc = docx.Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        logger.error(f"DOCX extraction error: {e}")
    return text

def extract_text_from_xlsx(file_path):
    text = ""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            text += f"\nSheet: {ws.title}\n"
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                text += row_text + "\n"
        wb.close()
    except Exception as e:
        logger.error(f"XLSX extraction error: {e}")
    return text

def _find_publication_date_in_xlsx(keywords: List[str]):
    """
    Scans XLSX files in UPLOAD_DIR for rows matching keywords,
    then attempts to extract a date.
    """
    files = []
    if not os.path.exists(UPLOAD_DIR):
        return None
        
    for fn in os.listdir(UPLOAD_DIR):
        if fn.lower().endswith(".xlsx") and ("journal" in fn.lower() or "journals" in fn.lower()):
            files.append(os.path.join(UPLOAD_DIR, fn))
            
    for fp in files:
        try:
            wb = load_workbook(filename=fp, read_only=True, data_only=True)
        except Exception:
            continue
            
        date_headers = {"date", "date of publication", "published on", "publication date"}
        
        for ws in wb.worksheets:
            header_map = {}
            try:
                rows_iter = ws.iter_rows(values_only=True)
            except Exception:
                continue
                
            first = True
            for row in rows_iter:
                vals = [str(c) for c in row if c is not None]
                row_text = " ".join(vals).lower()
                
                if first:
                    for i, c in enumerate(row):
                        s = str(c).lower() if c is not None else ""
                        if any(h in s for h in date_headers):
                            header_map["date_idx"] = i
                    first = False
                    
                # Check if all keywords are in the row
                if not all(k in row_text for k in keywords):
                    continue
                    
                date_candidates = []
                for c in row:
                    s = str(c) if c is not None else ""
                    # Try regex for date patterns
                    if re.search(r"\b20\d{2}-\d{1,2}-\d{1,2}\b", s):
                        date_candidates.append(s)
                    elif re.search(r"\b\d{1,2}/\d{1,2}/\d{4}\b", s):
                        date_candidates.append(s)
                    elif re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4}", s, re.IGNORECASE):
                        date_candidates.append(s)
                        
                if not date_candidates and "date_idx" in header_map:
                    idx = header_map["date_idx"]
                    if idx < len(row) and row[idx] is not None:
                        date_candidates.append(str(row[idx]))
                        
                if date_candidates:
                    wb.close()
                    return date_candidates[0]
        wb.close()
    return None

# --- App ---
app = FastAPI(title="MVSR RAG System")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    """On startup, initialize all background services in a separate thread."""
    logger.info("Initializing services. The server will be ready after this completes.")
    await asyncio.to_thread(initialize_services)
    logger.info("Service initialization complete. Server is now ready to accept requests.")
        
    if SECRET_KEY == "your-secret-key-here":
        logger.warning("Security warning: Using default SECRET_KEY. Please set a strong, unique key in your environment variables.")

@app.get("/health")
async def health_check(db = Depends(get_db)):
    db_status = "connected"
    if db is None:
        db_status = "disconnected"

    return {
        "status": "active", 
        "llm_health": check_llm_health(),
        "database_status": db_status
    }

@app.post("/token")
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: psycopg.AsyncConnection = Depends(get_db)):
    if not db:
        # Mock login for offline
        if form_data.username == "admin" and form_data.password == "admin123":
            return {"access_token": create_access_token(data={"sub": "admin"}), "token_type": "bearer"}
        raise HTTPException(status_code=400, detail="DB unavailable and invalid mock creds")

    async with db.cursor(row_factory=dict_row) as acur:
        await acur.execute("SELECT * FROM users WHERE username = %s", (form_data.username,))
        user = await acur.fetchone()
    
    if not user or not verify_password(form_data.password, user['password_hash']):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = create_access_token(data={"sub": user['username']})
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/api/query")
async def query_documents(
    query_request: Dict[str, Any],
    current_user: dict = Depends(get_current_user)
):
    query = query_request.get("query", "")
    history = query_request.get("history", [])  # Get history from request
    if not query:
        raise HTTPException(status_code=400, detail="Query is empty")

    # 1. Intent Analysis / Special Handling
    query_lower = query.lower()
    
    # Check for specific author/journal date query
    # E.g., "when did sharath kumar published journal of elictrical systems?"
    if "when" in query_lower and "publish" in query_lower:
        # Extract potential author and journal names (heuristic)
        # This is a simple heuristic, can be improved
        tokens = query_lower.split()
        # Remove common words
        stop = {"when", "did", "published", "publish", "in", "the", "of", "journal"}
        keywords = [t for t in tokens if t not in stop]
        
        # Try to find a date in XLSX
        date_found = _find_publication_date_in_xlsx(keywords) # Pass the whole list
        if date_found:
            return {"answer": f"According to the records, it was published on {date_found}.", "sources": ["journals.xlsx"]}

    # 2. Advanced RAG Retrieval
    context = ""
    sources = []
    
    if collection and retriever and context_optimizer:
        try:
            # 2.1. Initial Candidate Retrieval (from ChromaDB)
            initial_results = collection.query(
                query_texts=[query],
                n_results=20  # Retrieve more candidates for re-ranking
            )
            
            if initial_results and initial_results['documents']:
                # Reconstruct chunk dictionaries from ChromaDB results
                candidate_chunks = []
                for i, doc in enumerate(initial_results['documents'][0]):
                    meta = initial_results['metadatas'][0][i]
                    
                    # Parse JSON strings back into objects
                    if 'document_structure' in meta and isinstance(meta['document_structure'], str):
                        try:
                            meta['document_structure'] = json.loads(meta['document_structure'])
                        except json.JSONDecodeError:
                            # Handle cases where the string is not valid JSON
                            meta['document_structure'] = {}
                    if 'headers' in meta and isinstance(meta['headers'], str):
                        try:
                            meta['headers'] = json.loads(meta['headers'])
                        except json.JSONDecodeError:
                            meta['headers'] = []
                    if 'keywords' in meta and isinstance(meta['keywords'], str):
                        meta['keywords'] = [k.strip() for k in meta['keywords'].split(',')]


                    # The 'chunk_text' is the document itself
                    chunk_dict = {'chunk_text': doc, **meta}
                    candidate_chunks.append(chunk_dict)

                # 2.2. Re-ranking with Cross-Encoder
                reranked_chunks = retriever.rerank_chunks(query, candidate_chunks, top_k=10)

                # 2.3. Source Prioritization
                prioritized_chunks = context_optimizer.prioritize_sources(reranked_chunks)

                # 2.4. Diverse Retrieval
                diverse_chunks = retriever.diverse_retrieval(prioritized_chunks, max_chunks=7)
                
                # 2.5. Context Compression & Final Context Assembly
                window_size = retriever.dynamic_context_window(query)
                context = context_optimizer.compress_context(diverse_chunks, max_tokens=window_size)
                
                if diverse_chunks:
                    sources = list(set([chunk.get('source', 'unknown') for chunk in diverse_chunks]))

        except Exception as e:
            logger.error(f"Advanced RAG Query Error: {e}")
            context = "Error retrieving documents with advanced retrieval."

    # Fallback to simple retrieval if advanced fails or is disabled
    if (not context or "Error retrieving documents" in context) and collection:
        logger.info("Falling back to simple RAG retrieval.")
        try:
            results = collection.query(query_texts=[query], n_results=5)
            if results and results['documents']:
                context = "\n".join(results['documents'][0])
                if results['metadatas']:
                    sources = [m.get('source', 'unknown') for m in results['metadatas'][0]]
        except Exception as e:
            logger.error(f"Simple RAG Query Error: {e}")
            context = "Error retrieving documents."

    # Fallback to local file search if context is empty (Mock RAG)
    if not context and os.path.exists(UPLOAD_DIR):
        # Scan a few files
        for fn in os.listdir(UPLOAD_DIR)[:3]:
            if fn.endswith(".txt"):
                with open(os.path.join(UPLOAD_DIR, fn), 'r') as f:
                    context += f.read()[:1000] + "\n"

    # 3. Generate Answer
    answer = generate_answer_with_llm(context, query, history) # Pass history
    
    return {"answer": answer, "sources": list(set(sources))}

@app.get("/api/documents", response_model=List[Dict[str, Any]])
async def get_documents(
    current_user: dict = Depends(get_current_user), 
    db: psycopg.AsyncConnection = Depends(get_db)
):
    if not db:
        # Return mock data if DB is not available
        return [
            {"id": 1, "filename": "mock_paper.pdf", "department": "CSE", "category": "research", "uploaded_at": "2023-10-27T10:00:00Z"},
            {"id": 2, "filename": "mock_patent.docx", "department": "ECE", "category": "patent", "uploaded_at": "2023-10-26T12:00:00Z"},
        ]

    async with db.cursor(row_factory=dict_row) as acur:
        await acur.execute("SELECT id, filename, department, category, uploaded_at FROM documents ORDER BY uploaded_at DESC")
        documents = await acur.fetchall()
    return documents

@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    department: str = Form(...),
    category: str = Form(...),
    current_user: dict = Depends(get_current_user),
    db: psycopg.AsyncConnection = Depends(get_db)
):
    file_path = os.path.join(UPLOAD_DIR, file.filename)
    try:
        async with aiofiles.open(file_path, "wb") as buffer:
            content = await file.read()
            await buffer.write(content)
    except Exception as e:
        logger.error(f"Failed to write file {file.filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save file: {e}")

    # Extract text
    text = ""
    if file.filename.endswith(".pdf"):
        text = extract_text_from_pdf(file_path)
    elif file.filename.endswith(".xlsx"):
        text = extract_text_from_xlsx(file_path)
    elif file.filename.endswith(".docx"):
        text = extract_text_from_docx(file_path)
    elif file.filename.endswith(".txt"):
        try:
            async with aiofiles.open(file_path, "r", encoding="utf-8") as f:
                text = await f.read()
        except Exception as e:
            logger.error(f"Failed to read text file {file.filename}: {e}")
            # Decide if you want to stop or continue without text
            text = "" # Continue without text

    # Embed and Store
    if collection and text:
        try:
            chunk_data = create_semantic_chunks(text, document_id=file.filename)
            chunk_data = [chunk for chunk in chunk_data if chunk.get('quality_score', 0.3) >= 0.3]

            if chunk_data:
                documents = [chunk['chunk_text'] for chunk in chunk_data]
                metadatas = [chunk['metadata'] for chunk in chunk_data]
                ids = [chunk['embedding_id'] for chunk in chunk_data]

                current_timestamp = datetime.now(timezone.utc).isoformat()
                for i, meta in enumerate(metadatas):
                    meta['source'] = file.filename
                    meta['department'] = department
                    meta['created_at'] = current_timestamp
                    if 'page' not in meta:
                        meta['page'] = i
                    
                    if 'document_structure' in meta and isinstance(meta.get('document_structure'), dict):
                        meta['document_structure'] = json.dumps(meta['document_structure'])
                    if 'keywords' in meta and isinstance(meta.get('keywords'), list):
                        meta['keywords'] = ", ".join(meta['keywords'])
                    if 'headers' in meta and isinstance(meta.get('headers'), list):
                        meta['headers'] = json.dumps([h.get('text', '') for h in meta['headers']])

                collection.add(documents=documents, metadatas=metadatas, ids=ids)
        except Exception as e:
            logger.error(f"Embedding or ChromaDB storage failed for {file.filename}: {e}")


    # DB Record
    username = current_user.get('username', 'unknown_user')
    if db:
        try:
            async with db.cursor() as acur:
                await acur.execute(
                    """INSERT INTO documents (filename, file_path, department, category, uploaded_by)
                       VALUES (%s, %s, %s, %s, %s) RETURNING id""",
                    (file.filename, file_path, department, category, username)
                )
                await db.commit()
        except Exception as e:
            logger.error(f"DB Insert Error for {file.filename}: {e}")
            # Optionally rollback or handle error
            await db.rollback()
            
    return {"filename": file.filename, "status": "uploaded"}

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)

@app.delete("/api/documents/{document_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_document(
    document_id: int,
    current_user: dict = Depends(get_current_user),
    db: psycopg.AsyncConnection = Depends(get_db)
):
    if not db:
        raise HTTPException(status_code=503, detail="Database not available")

    # First, get the filename from the database
    async with db.cursor(row_factory=dict_row) as acur:
        await acur.execute("SELECT filename FROM documents WHERE id = %s", (document_id,))
        doc = await acur.fetchone()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        filename = doc['filename']

    # Delete from PostgreSQL
    async with db.cursor() as acur:
        await acur.execute("DELETE FROM documents WHERE id = %s", (document_id,))
        await db.commit()

    # Delete from ChromaDB
    if collection:
        # Find all chunks associated with the file
        results = collection.get(where={"source": filename})
        if results and results['ids']:
            collection.delete(ids=results['ids'])

    return
