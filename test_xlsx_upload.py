#!/usr/bin/env python3
import time
import requests
import os
from openpyxl import Workbook
BACKEND_URL = os.environ.get('BACKEND_URL', 'http://localhost:8001')

# Wait for backend to start
print('Waiting for backend to start...')
backend_running = False
for i in range(15):
    try:
        response = requests.get(f'{BACKEND_URL}/docs', timeout=5)
        if response.status_code == 200:
            print('Backend is running!')
            backend_running = True
            break
    except:
        pass
    time.sleep(2)
    print(f'Attempt {i+1}/15...')
if not backend_running:
    print('Backend failed to start, will use local XLSX parsing fallback.')

headers = {}
if backend_running:
    print('Testing login...')
    login_data = {'username': 'admin', 'password': 'admin123'}
    response = requests.post(f'{BACKEND_URL}/token', data=login_data)
    print(f'Login status: {response.status_code}')
    if response.status_code != 200:
        print(f'Login failed: {response.text}')
    else:
        token = response.json()['access_token']
        print('Login successful')
        headers = {'Authorization': f'Bearer {token}'}
        response = requests.get(f'{BACKEND_URL}/api/documents', headers=headers)
        print(f'Documents status: {response.status_code}')
        docs = response.json()
        print(f'Current documents: {len(docs.get("documents", []))}')

# Create a test XLSX file
print('Creating test XLSX file...')
wb = Workbook()
ws = wb.active
ws.title = "Test Data"
ws.append(["Name", "Department", "Score"])
ws.append(["Alice", "CSE", "95"])
ws.append(["Bob", "ECE", "87"])
ws.append(["Charlie", "MECH", "92"])
test_file = "test_data.xlsx"
wb.save(test_file)
wb.close()

# Upload the XLSX file
if backend_running and headers.get('Authorization'):
    print('Uploading XLSX file...')
    with open(test_file, 'rb') as f:
        files = {'file': (test_file, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'department': 'CSE', 'category': 'research'}
        response = requests.post(f'{BACKEND_URL}/api/upload', files=files, data=data, headers=headers)
    print(f'Upload status: {response.status_code}')
    print(f'Upload response: {response.text}')
    if response.status_code == 200:
        if os.path.exists(f'./uploads/{test_file}'):
            print('File saved successfully')
        else:
            print('File not found in uploads directory')
        print('Testing query...')
        query_data = {'query': 'What is Alice score?', 'top_k': 5}
        response = requests.post(f'{BACKEND_URL}/api/query', json=query_data, headers=headers)
        print(f'Query status: {response.status_code}')
        if response.status_code == 200:
            result = response.json()
            print(f'Query response: {result.get("response", "")[:200]}...')
            print(f'Sources: {len(result.get("sources", []))}')

# Show chunks from journals XLSX
journals_filename = 'Journals-Conferences-Books-Book Chapters 2024-25 (1).xlsx'
try:
    enc = requests.utils.requote_uri(journals_filename)
    url = f'{BACKEND_URL}/api/chunks?filename={enc}'
    print('Fetching chunks from journals XLSX...')
    r = requests.get(url, headers=headers, timeout=30)
    print(f'Chunks fetch status: {r.status_code}')
    if r.status_code == 200:
        data = r.json()
        print(f'Chunk count: {data.get("count")}')
        for item in data.get('chunks', [])[:8]:
            idx = item.get('chunk_index')
            txt = (item.get('chunk_text') or '')[:300].replace('\n', ' ')
            print(f'CHUNK {idx}: {txt}')
    else:
        print(r.text)
except Exception as e:
    print('Failed to fetch chunks via API:', e)
    try:
        print('Falling back to local XLSX parsing...')
        from openpyxl import load_workbook
        fp = os.path.join('uploads', journals_filename)
        if not os.path.exists(fp):
            print('Journals XLSX not found at', fp)
        else:
            wb = load_workbook(filename=fp, read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets:
                parts.append(f'Sheet: {ws.title}')
                for row in ws.iter_rows(values_only=True):
                    rv = [str(c) for c in row if c is not None]
                    if rv:
                        parts.append(' '.join(rv))
            wb.close()
            text = '\n'.join(parts)
            chunk_size = 500
            overlap = 50
            chunks = []
            start = 0
            while start < len(text):
                end = start + chunk_size
                chunks.append(text[start:end])
                start = end - overlap
            print(f'Local chunk count: {len(chunks)}')
            for i, c in enumerate(chunks[:8]):
                safe = (c[:300]).replace('\n', ' ')
                print(f'CHUNK {i}: {safe}')
    except Exception as e2:
        print('Local XLSX parsing failed:', e2)

# Clean up
if os.path.exists(test_file):
    os.remove(test_file)
    print('Cleaned up test file')
